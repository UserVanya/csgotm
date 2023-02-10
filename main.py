import datetime
import auxiliarry as aux
import time
import threading
import PySimpleGUI as sg
import sys
import openpyxl

def get_prev_data():
    wb = openpyxl.open('deals.xlsx')
    ws = wb.active
    deals = []
    for i in range(1, ws.max_row):
        deals.append([ws.cell(i, 1).value, float(ws.cell(i, 2).value), ws.cell(i, 3).value])
    return deals


def scanning_thread(total_items):
    previous_items = aux.scan_new_items()
    total_items += previous_items
    wb = openpyxl.open('deals.xlsx')
    ws = wb.active
    start = datetime.datetime.now()
    while True:
        time.sleep(5)
        new_items = aux.scan_new_items()
        to_add_items = aux.get_diff(new_items, previous_items)
        if to_add_items is not None:
            total_items += to_add_items
            previous_items += to_add_items
        previous_items = aux.get_last_items(previous_items)
        for item in to_add_items:
            ws.append(item)
        if (datetime.datetime.now() - start).total_seconds() > 60:
            wb.save('deals.xlsx')
            start = datetime.datetime.now()


def observe_item_with_adds(name, deals):
    try:
        deals_local = [deal for deal in deals if deal[0] == name]
        prices = [deal[1] for deal in deals_local]
        name_found, cost, amount = aux.get_steam_item_info(name)
        average_price = sum(prices) / len(prices)
        layout = [[sg.Text(f"Коэффициент: {round(cost / average_price, 2)} ({round(cost/max(prices), 2)} - {round(cost/min(prices), 2)})")],
                  [sg.Button("STEAM"), sg.Text(f"STEAM: от {cost} р.  Количество: {amount}")],
                  [sg.Button('CSGOTM'), sg.Text(f"Средняя цена: {average_price} р. ({min(prices)} р. - {max(prices)} р.)")],
                  [sg.Text('Последние сделки на CSGOTM:')],
                  [sg.Listbox(values=deals_local, size=(80, 10))]]
        window = sg.Window(name_found, layout)
        event = None
        while event is None:
            event, values = window.read()
            if event in (None, 'Exit', 'Отмена'):
                window.close()
                return
            if event in ('CSGOTM', 'STEAM'):
                aux.open_link(event, name)
                event = None
    except:
        print(sys.exc_info())


start_time = datetime.datetime.strptime(datetime.datetime.today().strftime('%d %B, %Y %Hh'), '%d %B, %Y %Hh')
deals = get_prev_data()
#print(deals)
with_adds = []
no_adds = []
x = threading.Thread(target=scanning_thread, args=(deals,), daemon=True)
x.start()
end_date = aux.cur_moment
layout = [[sg.Text('С'), sg.Button('начала работы программы', key='start_moment', enable_events=True),
           sg.Text('до'), sg.Button('текущего момента', key='end_moment', enable_events=True)],
          [sg.Text('Цена 1 вещи от:'), sg.InputText(0, key='min_price', size=(7, 20)),
           sg.Text('до'), sg.InputText(100000, key='max_price', size=(7, 20)), sg.Text('рублей')],
          [sg.Text('Количество сделок от:'), sg.InputCombo(list(range(0, 99)), key='min_buys', default_value=0),
           sg.Text('до'), sg.InputCombo(list(range(0, 99)), key='max_buys', default_value=99)],
          [sg.Text('Отображаемое кол-во элементов:'), sg.InputText(30, key='amount', size=(3, 20))],
          [sg.InputText(size=(63, 20), key='find_in'), sg.Button('Стереть', key='refine1'),
           sg.Button('Выбрать', key='find')],
          [sg.Listbox(values=with_adds, key='with_adds', size=(80, 20), enable_events=True),
           sg.Listbox(values=no_adds, key='no_adds', size=(80, 20), enable_events=True)],
          [sg.Button('Обновить'), sg.Button('Выход')]]
window = sg.Window('Топ ликвидности', layout)
event = None
while event is None:
    event, values = window.read()
    if event in (None, 'Exit', 'Отмена'):
        window.close()
        sys.exit()
    if event == 'Обновить':
        with_adds = aux.get_best(deals, int(values['amount']),
                                 (int(values['min_price']), int(values['max_price'])),
                                 (start_time, end_date), (int(values['min_buys']), int(values['max_buys'])),
                                 values['find_in'])
        no_adds = aux.get_best(aux.get_items_no_adds(deals), int(values['amount']),
                               (int(values['min_price']), int(values['max_price'])),
                               (start_time, end_date), (int(values['min_buys']), int(values['max_buys'])),
                               values['find_in'])
        window['with_adds'].Update(values=with_adds)
        window['no_adds'].Update(values=no_adds)
        #print(window['with_adds'].get_list_values())
        #print(values)
        event = None
    elif event == 'start_moment':
        start_time, start_date_string = aux.choose_beginning_time(start_time, window['start_moment'].get_text())
        window['start_moment'].Update(start_date_string)
        event = None
    elif event == 'end_moment':
        end_date, end_date_string = aux.choose_ending_time(window['end_moment'].get_text())
        window['end_moment'].Update(end_date_string)
        event = None
    elif event == 'no_adds':
        window['find_in'].Update(values['no_adds'][0][0])
        event = None
    elif event == 'with_adds':
        window['find_in'].Update(values['with_adds'][0][0])
        event = None
    elif event == 'find':
        observe_item_with_adds(values['find_in'], deals)
        event = None
    elif event == 'refine1':
        window['find_in'].Update('')
        event = None
    else:
        print('no event found:', event)
